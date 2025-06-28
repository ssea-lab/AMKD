import os
import sys
import argparse
import timm
import torch
import numpy as np
import pandas as pd
import openpyxl
import re, ast
import torch.nn.functional as F
import models.resnet as resnet
import models.resnet1 as resnet1
import models.convnext as convnext
import models.swin_transformer as swin_transformer
from torch.utils.data import DataLoader
from dataset import FrameDataSetLabel
from util import seeding, create_directory
from sklearn.metrics import roc_auc_score
from sklearn.metrics import confusion_matrix, accuracy_score
from decimal import Decimal
from models.resnet import *
from models.resnet1 import *
from models.convnext import *
from models.swin_transformer import *


def parse_option():
    parser = argparse.ArgumentParser('the parameters for testing the model performance')

    parser.add_argument('--model_s', default='resnet_little', type=str, choices=['resnet_little', 'convnext_little',
                                                                                 'resnet_lite', 'convnext_lite',
                                                                                 'swin_little', 'swin_lite'],
                        help='the student model we used.')
    parser.add_argument('--model_t', default='resnet18', type=str, choices=['resnet18', 'convnext_pico', 'swin_tiny'],
                        help='the teacher model we used.')
    parser.add_argument('--num_class', default=5, type=int, help="class num")
    parser.add_argument('--model_name', default='', type=str, help="just for placeholder")
    parser.add_argument('--cam_type', type=str, default='gradcam', choices=['gradcam', 'eigen_cam'])

    # parameters for data
    parser.add_argument('--crop_frame_height', type=int, default=600, help='the frame height we used.')
    parser.add_argument('--frame_height', type=int, default=512, help='the frame height we used during training.')
    parser.add_argument('--frame_width', type=int, default=1024, help='the frame width we used during training.')

    # folder num，以便用于内部数据集的交叉验证
    parser.add_argument('--fold_num', type=int, default=0, choices=[0, 1, 2, 3, 4],
                        help='the fold num we used for cross validation.')

    # model parameters initialization pattern
    parser.add_argument('--pre_train', action='store_true', help="weight initialized by the weight pretrained from "
                                                                 "imageNet")
    parser.add_argument('--use_class_weight', action='store_true', help="whether do linear probing or not")
    parser.add_argument('--use_label_smoothing', action='store_true', help="whether do linear probing or not")
    parser.add_argument('--label_smoothing', type=float, default=0.1, help="whether do linear probing or not")

    # 这个地方用于证明大规模无标签数据用于蒸馏时可以很好的提高模型的性能
    parser.add_argument('--use_unlabel_data', action='store_true', help="whether use unlabel data for distillation")
    parser.add_argument('--use_pselabel', action='store_true',
                        help="whether use psedeuo label data for unlabeled distillation")
    parser.add_argument('--use_label_data', action='store_true', help="whether use label data for distillation")

    # distillation
    parser.add_argument('--distill', type=str, default='amkd')
    # configuration for the weight of different loss function
    parser.add_argument('--am_weight', type=float, default=1.0, help='weight of instance RKD loss')
    parser.add_argument('--kl_weight', type=float, default=1.0, help='weight of ce loss')
    parser.add_argument('--ce_weight', type=float, default=1.0, help='weight of ce loss')

    # whether use the loss function
    parser.add_argument('--use_am', action='store_true', help="whether use the irkd loss")
    parser.add_argument('--use_kl', action='store_true', help="whether use the kl loss")
    parser.add_argument('--use_ce', action='store_true', help="whether use the ce loss")

    # test_mode configuration，方便查找预训练权重的路径
    parser.add_argument('--test_pattern', type=str, default="supervised_train")

    parser.add_argument('--weight_decay', default=0.6, type=float, help="weight_decay")
    parser.add_argument('--weight_decay_end', type=float, default=0.4, help="""Final value of the
           weight decay. We use a cosine schedule for WD.""")

    parser.add_argument('--test_file', default='internal_test',
                        choices=['internal_test', 'tiff_huaxi', 'tiff_xiangya'],
                        help="the test file from different center", type=str)

    parser.add_argument('--test_kind', default='frame',
                        choices=['frame', 'volume'],
                        help="the test kind for different center", type=str)

    parser.add_argument('--batch_size', type=int, default=1, help="training batch size")
    parser.add_argument('--num_workers', type=int, default=0, help="data loader thread")
    parser.add_argument('--gpu', type=int, default=0, help='the gpu number will be used')
    parser.add_argument('--checkpoint', type=str, default='checkpoint',
                        help='the directory to load the model trained weights.')
    parser.add_argument('--seed', default=42, type=int)
    parser.add_argument('--result_dir', default='result', type=str, help='the directory of the model testing '
                                                                         'performance')

    args = parser.parse_args()
    return args


def path_determine(args):
    # 获取学生模型的权重存储路径
    assert args.use_am or args.use_kl or args.use_ce, '所有的损失函数不能同时为False'
    if not args.use_am:
        args.am_weight = 0
    if not args.use_kl:
        args.kl_weight = 0
    if not args.use_ce:
        args.ce_weight = 0
    args.save_postfix = 'S:{}_T:{}_distill:{}' \
                        '_am:{}_kl:{}_ce:{}'.format(args.model_s,
                                                    args.model_t, args.distill,
                                                    args.am_weight,
                                                    args.kl_weight, args.ce_weight)

    if args.use_unlabel_data and args.use_label_data:
        args.label_data = 'all_data_pselabel' if args.use_pselabel else 'all_data'
    else:
        args.label_data = 'unlabel_data' if args.use_unlabel_data else 'label_data'

    args.save_dir = os.path.join('checkpoint', 'distill', args.label_data, args.cam_type, args.save_postfix, str(args.fold_num))
    args.directory_path = os.path.join('distill', args.label_data, args.cam_type, args.save_postfix)
    args.weight_path = os.path.join(args.save_dir, 'checkpoint.pth')
    return args


def get_model(args, distill=False, teacher=False):
    model_name = args.model_name
    if distill:
        model_name = args.model_t if teacher else args.model_s
    pre_train = args.pre_train
    num_class = args.num_class
    model = None
    # 定义模型
    if model_name == 'resnet18':
        model = getattr(resnet, model_name)(pretrained=False, num_classes=num_class)
    elif model_name in ['resnet_little', 'resnet_lite']:
        model = getattr(resnet1, model_name)(num_classes=num_class)
    elif model_name in ['convnext_pico', 'convnext_little', 'convnext_lite']:
        model = getattr(convnext, model_name)(pretrained=False, num_classes=num_class)
    else:
        model = getattr(swin_transformer, model_name)(pretrained=False, num_classes=num_class)
    return model


def load_weight(model, weight_path):
    # 加载模型模型权重
    print('==> loading teacher model')
    checkpoint = torch.load(weight_path, map_location='cpu')
    model_dict = checkpoint['state_dict']
    state = model.load_state_dict(model_dict, strict=False)
    print('loading checkpoint from {}'.format(weight_path))
    print(state)
    return model


def worker_frame(args):
    args = path_determine(args)
    # 定义帧的测试数据集
    test_set = FrameDataSetLabel(args, pattern='valid', augmentation='valid_or_test')
    test_loader = DataLoader(dataset=test_set,
                             batch_size=1,
                             pin_memory=True,
                             shuffle=False)

    # 定义模型
    model = get_model(args, distill=True)
    model = load_weight(model, args.weight_path)
    device = torch.device('cuda:' + str(args.gpu))
    model = model.to(device)

    result_dir = os.path.join(args.result_dir, args.test_kind, args.test_file, args.directory_path)
    create_directory(result_dir)
    test_result_path = os.path.join(result_dir, 'test_result.xlsx')
    total_result_path = os.path.join(result_dir, 'total_result.xlsx')

    test_frame(test_loader, model, test_result_path, args.fold_num)
    get_metrics_from_file(test_result_path, total_result_path, args.fold_num)


def test_frame(test_loader, model, result_path, fold_num):
    # fold_num: int
    model.eval()
    device = next(model.parameters()).device
    # 定义所要输出的文件的表头
    if os.path.isfile(result_path):
        # 加载现有的Excel工作簿
        workbook = openpyxl.load_workbook(result_path)
        # 根据fold_num创建对应的worksheet
        worksheet = workbook.create_sheet('fold_{}'.format(fold_num))
    else:
        # 创建一个excel表格用于存储预测结果
        workbook = openpyxl.Workbook()
        # 获取当前活跃的worksheet，默认就是第一个worksheet
        worksheet = workbook.active
        worksheet.title = 'fold_{}'.format(fold_num)

    header = ['图片路径', '五分类预测概率', '五分类预测标签', '标签']

    # 添加结果excel文件的表头
    worksheet.append(header)

    with torch.no_grad():
        for j, data in enumerate(test_loader):
            images, labels, _, _, img_paths = data
            batch_size = len(labels)
            images = images.to(device, non_blocking=True)
            outputs = model(images)
            outputs = F.softmax(outputs, dim=1)

            # 接下来我们要把模型预测的结果和真实的标签写到结果文件中
            outputs = outputs.cpu().numpy()
            labels = labels.numpy()

            output_labels = outputs.argmax(axis=1)  # [1]
            img_path = img_paths[0]
            outputs = outputs[0].tolist()

            # 输出结果的类别格式
            # image_path, 5分类的预测概率, 5分类的预测标签, 5分类的真实标签
            outputs = ' '.join(list(map(str, outputs)))

            output_labels = output_labels[0]
            labels = labels[0]

            content = [img_path, outputs, output_labels, labels]
            worksheet.append(content)

    workbook.save(filename=result_path)


def get_metrics_from_file(test_result_file, total_result_file, fold_num):
    # fold_num: int
    # 所要计算的指标包括五分类准确率(five class accuracy)，五分类的mac_croF1, 二分类准确率(binary class accuracy),
    # 灵敏度(sensitivity), 特异性(specificity), PPV, NPV, AUC
    # test_result_file 输出结果的类别格式
    # ['图片路径', '5分类预测概率', '5分类预测标签', '标签']
    # 定义所要输出的文件的表头
    if os.path.isfile(total_result_file):
        # 加载现有的Excel工作簿
        workbook = openpyxl.load_workbook(total_result_file)
        # 根据fold_num创建对应的worksheet
        worksheet = workbook.create_sheet('fold_{}'.format(fold_num))
    else:
        # 创建一个excel表格用于存储预测结果
        workbook = openpyxl.Workbook()
        # 获取当前活跃的worksheet，默认就是第一个worksheet
        worksheet = workbook.active
        worksheet.title = 'fold_{}'.format(fold_num)

    # 定义所要输出的文件的表头
    header = ['五分类准确率', '二分类准确率', '灵敏度', '特异性', 'PPV', 'NPV', 'AUC',
              '二分类混淆矩阵', '五分类混淆矩阵']
    worksheet.append(header)

    five_class_pre = []
    binary_class_pre = []
    five_class_labels = []
    binary_class_labels = []
    binary_class_probabilities = []
    # 读取结果的excel文件
    df = pd.read_excel(io=test_result_file, sheet_name='fold_{}'.format(fold_num))
    print(f"计算第{fold_num}结果")
    for idx, row in df.iterrows():
        # ['图片路径', '5分类预测概率', '5分类预测标签', '标签']
        img_path = row[0]
        five_probs = list(map(float, row[1].split(' ')))
        five_pre = int(row[2])
        five_label = int(row[3])
        # 根据5分类的预测概率推导出2分类的预测标签
        pos_p, neg_p = five_probs[3] + five_probs[4], five_probs[0] + five_probs[1] + five_probs[2]
        binary_pre = 1 if pos_p > neg_p else 0

        binary_label = 1 if five_label > 2 else 0

        five_class_pre.append(five_pre)
        binary_class_pre.append(binary_pre)
        five_class_labels.append(five_label)
        binary_class_labels.append(binary_label)
        binary_class_probabilities.append(pos_p)

    five_class_pre = np.array(five_class_pre)
    binary_class_pre = np.array(binary_class_pre)
    five_class_labels = np.array(five_class_labels)
    binary_class_labels = np.array(binary_class_labels)
    binary_class_probabilities = np.array(binary_class_probabilities)

    # 利用sklearn的接口计算得到五分类的准确率和二分类的准确率
    five_accracy = accuracy_score(y_true=five_class_labels, y_pred=five_class_pre)
    # 利用sklearn的接口计算得到混淆矩阵
    five_confusion_matrix = confusion_matrix(y_true=five_class_labels,
                                             y_pred=five_class_pre)

    # 接下来利用二分类的混淆矩阵计算各种指标，包括灵敏度，特异性，PPV, NPV
    # metrics = [binary_accracy, sensitivity, specifity, ppv, npv, binary_confusion_matrix]
    metrics = cal_bin_metrics(y_true=binary_class_labels, y_pred=binary_class_pre)
    binary_accracy, sensitivity, specifity, ppv, npv, binary_confusion_matrix = metrics[0], metrics[1], metrics[2], \
        metrics[3], metrics[4], metrics[5]

    # 接下来计算二分类的auc的值
    auc = roc_auc_score(y_true=binary_class_labels, y_score=binary_class_probabilities)
    # 接下来将输出结果规范化
    metrics1 = list(map(norm_value, [five_accracy, binary_accracy, sensitivity, specifity, ppv, npv]))
    auc = norm_value(auc, percent=False)
    metric2 = list(map(str, [binary_confusion_matrix, five_confusion_matrix]))
    metrics = [*metrics1, auc, *metric2]
    worksheet.append(metrics)
    workbook.save(filename=total_result_file)

    # 如果fold_num=4,则要新创建一个子表用于计算五折的结果的均值和方差
    if fold_num == 4:
        result_list = []
        for i in range(fold_num):
            result_list.append(pd.read_excel(total_result_file, sheet_name=f'fold_{i}').iloc[:1])
        df = pd.concat(result_list)
        df.reset_index(drop=True, inplace=True)

        total_list = []
        # 由于二分类混淆矩阵和五分类混淆矩阵是以字符串的方式存储的，所以不能直接拿来求均值和方差
        for column in df.columns:
            if column not in ['二分类混淆矩阵', '五分类混淆矩阵']:
                mean = norm_value(float(df[column].mean()), percent=False)
                std = norm_value(float(df[column].std()), percent=False)
                total_list.append(f'{mean}±{std}')
            else:
                conf_matrix_list = []
                conf_matrix = df[column]  # Series of string
                for conf_matrix_str in conf_matrix:
                    conf_matrix_list.append(npstr2array(conf_matrix_str))

                total_arr = np.zeros(conf_matrix_list[0].shape)
                for arr in conf_matrix_list:
                    total_arr = total_arr + arr
                avg_arr = total_arr / len(conf_matrix_list)
                avg_arr = avg_arr.astype(int)
                total_list.append(str(avg_arr))

        worksheet = workbook.create_sheet('fold_total')
        worksheet.append(header)
        worksheet.append(total_list)
        workbook.save(filename=total_result_file)


def norm_value(number, percent=True):
    if percent:
        number = Decimal(number).quantize(Decimal('0.0001'), rounding='ROUND_HALF_UP') * 100
    else:
        number = Decimal(number).quantize(Decimal('0.01'), rounding='ROUND_HALF_UP')
    return number


def npstr2array(npstr):
    npstr = re.sub("\[\s+", '[', npstr)
    npstr = re.sub('\s+', ',', npstr)
    arr = np.array(ast.literal_eval(npstr))
    return arr


def cal_bin_metrics(y_true, y_pred):
    """
    the function is used to calculate the corresponding metrics for binary classification,
    including 二分类准确率，灵敏度，特异性，PPV, NPV, AUC
    :param y_true:
    :param y_pred:
    :return:
    """
    # 首先判断一下y_true, y_pred的维度，再进行相应的操作
    if y_pred.ndim == 2:
        y_pred = np.argmax(y_pred, axis=1)
    binary_accracy = accuracy_score(y_true=y_true, y_pred=y_pred)
    # 利用sklearn的接口计算得到混淆矩阵
    binary_confusion_matrix = confusion_matrix(y_true=y_true, y_pred=y_pred)
    # print(y_true)
    # print(y_pred)
    # print(binary_confusion_matrix)
    # 接下来利用二分类的混淆矩阵计算各种指标，包括灵敏度，特异性，PPV, NPV
    tp = binary_confusion_matrix[1][1]
    fn = binary_confusion_matrix[1][0]
    fp = binary_confusion_matrix[0][1]
    tn = binary_confusion_matrix[0][0]
    sensitivity = tp / (tp + fn)
    specifity = tn / (tn + fp)
    ppv = tp / (tp + fp)
    npv = tn / (tn + fn)
    metrics = [binary_accracy, sensitivity, specifity, ppv, npv, binary_confusion_matrix]
    return metrics


if __name__ == '__main__':
    args = parse_option()
    # 设置随机数种子
    seeding(args.seed)
    for fold_num in [0, 1, 2, 3, 4]:
        # for fold_num in [0]:
        args.fold_num = fold_num
        worker_frame(args)
