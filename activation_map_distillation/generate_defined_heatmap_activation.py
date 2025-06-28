"""
这个脚本主要是根据分割掩码以及边框盒标记的区域生成自定义的热力图激活，
这个热力图激活用来指导分类时模型应激活的区域。
"""
import math
import os

import numpy as np
import cv2
import torch
import time
import matplotlib.pyplot as plt
import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from collections import defaultdict
from collections.abc import Iterable
from copy import deepcopy

true_label_sequence = ['background', '背景', '上皮', '间隙', '保护套', '凸起', '囊肿', '基质']


def find_closest_point(points_total, target_points_total):
    print(len(target_points_total))
    """
    寻找列表中离给定点最近的点和距离。

    :param points: 一个包含多个二维点的列表，例如 [(x1, y1), (x2, y2), ...]， 长度为m
    :param target_points: 给定的点的列表，格式为  [(x1, y1), (x2, y2), ...], 长度为n
    :return: 最近点和距离
    """
    num_target_points = len(target_points_total)
    max_points = 10000
    iters = int(math.ceil(num_target_points / max_points))
    device = torch.device("cuda:1" if torch.cuda.is_available() else "cpu")
    total_distance = torch.empty(0, len(points_total)).to(device)

    for i in range(iters):
        target_points = target_points_total[i * max_points: (i + 1) * max_points]
        # 将列表转换为numpy数组以进行高效计算
        points = np.array(points_total)
        m = points.shape[0]
        target_points = np.array(target_points)
        n = target_points.shape[0]

        # 将numpy array转化为Tensor张量
        points = torch.from_numpy(points)
        target_points = torch.from_numpy(target_points)
        # 将张量移动到GPU上去
        points = points.to(device)
        target_points = target_points.to(device)

        # 将points进行扩展维度
        points = points.unsqueeze(0).expand(n, m, 2).reshape(n * m, 2)
        # 将target_points进行扩展维度
        target_points = target_points.unsqueeze(1).expand(n, m, 2).reshape(n * m, 2)
        # 计算所有点到目标点的欧式距离
        distances = torch.sqrt(torch.sum((points - target_points) ** 2, dim=1))  # [n*m]
        distances = distances.reshape(n, m)
        total_distance = torch.cat((total_distance, distances), dim=0)
    # 将distances从GPU移动到CPU上去，并转化为ndarray
    distances = total_distance.cpu().numpy()

    # 按行求最小值
    distances_min_row = np.min(distances, axis=1)
    max_dis = np.max(distances_min_row)
    min_dis = np.min(distances_min_row)

    return min_dis, max_dis, distances_min_row


def linear_function_through_points(point1, point2):
    # 解包输入的两个点
    x1, y1 = point1
    x2, y2 = point2

    # 计算斜率
    m = (y2 - y1) / (x2 - x1)

    # 计算截距
    b = y1 - m * x1

    # 定义并返回线性函数
    def linear_function(x):
        return m * x + b

    return linear_function


def get_hm_activation_according_to_distances(duibi_position_list, now_position_list, upper_bound, lower_bound):
    """
    这个函数用来根据两个列表中的点，生成自定义的热力图激活
    :param duibi_position_list:
    :param now_position_list:
    :param upper_bound:
    :param lower_bound:
    :return:
    """
    if len(duibi_position_list) == 0:
        # 如果要对比的列表为空，则直接返回一个全为0.8的激活列表
        activation_list = [0.8] * len(now_position_list)
    else:
        now_duibi_dis_min, now_duibi_dis_max, now_duibi_distance_list = find_closest_point(duibi_position_list,
                                                                                           now_position_list)
        linear_func = linear_function_through_points((now_duibi_dis_min, upper_bound),
                                                     (now_duibi_dis_max, lower_bound))

        # 根据这个线性函数，计算每一个需求位置激活值
        activation_list = []
        for i in range(len(now_duibi_distance_list)):
            activation_list.append(linear_func(now_duibi_distance_list[i]))

    return activation_list


def find_pos_np(matrix, value):
    # 使用np.where找出值为mask的元素位置
    pos_list = list(zip(*np.where(matrix == value)))

    def remove_zero_coordinates(coordinates):
        # 移除横坐标或纵坐标为0的元组
        return [(x, y) for x, y in coordinates if x != 0 and y != 0]

    return remove_zero_coordinates(pos_list)


def get_coordinates(matrix):
    # 获取行索引和列索引
    row_indices, col_indices = np.indices(matrix.shape)
    row_indices = row_indices[1:-1, :]
    col_indices = col_indices[:, 1:-1]
    # 将行列索引配对并转换为列表
    return list(set(list(zip(row_indices.ravel(), col_indices.ravel()))))


def show_cam_on_image(img: np.ndarray,
                      mask: np.ndarray,
                      use_rgb: bool = False,
                      colormap: int = cv2.COLORMAP_JET) -> np.ndarray:
    """ This function overlays the cam mask on the image as an heatmap.
    By default the heatmap is in BGR format.

    :param img: The base image in RGB or BGR format.
    :param mask: The cam mask.
    :param use_rgb: Whether to use an RGB or BGR heatmap, this should be set to True if 'img' is in RGB format.
    :param colormap: The OpenCV colormap to be used.
    :returns: The default image with the cam overlay.
    """

    heatmap = cv2.applyColorMap(np.uint8(255 * mask), colormap)
    if use_rgb:
        heatmap = cv2.cvtColor(heatmap, cv2.COLOR_BGR2RGB)
    heatmap = np.float32(heatmap) / 255

    if np.max(img) > 1:
        raise Exception(
            "The input image should np.float32 in the range [0, 1]")

    cam = heatmap + img
    cam = cam / np.max(cam)
    return np.uint8(255 * cam)


def fuse_gray_scale_activations(image, position_list, dis_activations):
    grayscales = np.array([image[pos[0], pos[1]] for pos in position_list])
    gray_activations = np.minimum(grayscales / grayscales.mean(), 1)
    activations = [dis_activations[i] * gray_activations[i] for i in
                   range(len(position_list))]
    return activations


def generate_zhengchang_heatmap_activation(image, mask):
    """
        这个函数用来根据宫颈炎图片的分割掩码生成自定义的热力图激活
        :param image: 3d array, [H, W, 3]
        :param mask: 2d array, [H, W]
        :return: 热力图激活叠加原图的3d array, [H, W, 3]，可以直接展示
        """
    jidimo_position_list = []  # list of array, 如[(120, 234), (500, 600)]
    # 首先遍历mask,以便于找到基底膜所在的像素位置，并且将每个位置保存在一个列表中
    H, W = mask.shape[0], mask.shape[1]

    # 个人觉得边界问题可以忽略，考虑边界问题只会增加代码的复杂度，因为边界只有最左边一列，最右边
    # 一列，最上边一行和最下边一行，所占据的像素数量较少，因此在手动创建热力图激活值时可以先不予考虑，
    # 等[(H-1)*(W-1)]的热力图激活值都生成后，再考虑边界问题，通过补齐或插值的方式生成[H*W]的热力图激活值
    for i in range(1, H - 1):
        for j in range(1, W - 1):
            # 分两种情况进行讨论，
            # 1. 当前像素值上皮时，当其左、右、下的像素为基质时，那么该像素可以看做是基底膜
            if mask[i, j] == 2:
                if mask[i, j - 1] == 7 or mask[i, j + 1] == 7 or mask[i + 1, j] == 7:
                    jidimo_position_list.append((i, j))

            # 2. 当当前像素值为基质时，当其左、右、上的像素为上皮时，那么该像素可以看做是基底膜
            if mask[i, j] == 7:
                if mask[i, j - 1] == 2 or mask[i, j + 1] == 2 or mask[i - 1, j] == 2:
                    jidimo_position_list.append((i, j))

    jidimo_position_list = list(set(jidimo_position_list))
    # 这里就存在一个问题，如果一个图片中因为采样问题导致没有基底膜像素怎么办，即jidimo_position_list为空，这种情况下要单独处理
    if len(jidimo_position_list) == 0:
        jidimo_position_list = []
        jidimo_activation_list = []
        print('基底膜元素个数为0')
    else:
        # 基底膜位置处的像素激活值设置为1
        jidimo_activation = 1.0
        jidimo_activation_list = [jidimo_activation for i in range(len(jidimo_position_list))]
        print('基底膜激活值已经计算完成')

    # 定义上皮位置列表，用来保存
    shangpi_position_list = find_pos_np(mask, 2)
    # 这里就存在一个问题，如果一个图片中因为采样问题导致没有上皮像素怎么办，即shangpi_position_list为空，这种情况下要单独处理
    if len(shangpi_position_list) == 0:
        shangpi_position_list = []
        shangpi_activations = []
        print('上皮元素个数为0')
    else:
        # 去除上皮位置列表与基底膜位置列表中重复的元素
        shangpi_position_list = list(set(shangpi_position_list) - set(jidimo_position_list))
        # 定义上皮激活值的上界upper_bound和下界lower_bound
        sp_activation_upper_bound = 1.0
        sp_activation_lower_bound = 0.5
        shangpi_dis_activations = get_hm_activation_according_to_distances(jidimo_position_list, shangpi_position_list,
                                                                           sp_activation_upper_bound,
                                                                           sp_activation_lower_bound)

        # 定义滤波器的大小
        kernel_size = (7, 7)
        # 对数组进行平均滤波
        image = image[:, :, 0]  # 取一个channel变为灰度图
        image = cv2.blur(image, kernel_size)
        # 融合上皮的灰度值重新计算激活值
        shangpi_activations = fuse_gray_scale_activations(image, shangpi_position_list, shangpi_dis_activations)

        print('上皮激活值已经计算完成')
    # 到目前为止，每一个上皮所在的像素位置与对应热力图激活值都已经计算出来了

    # 接下来计算基质对应的每一个像素的位置的热力图激活值
    # 定义基质位置列表，用来保存
    jizhi_position_list = find_pos_np(mask, 7)
    # 这里就存在一个问题，如果一个图片中因为采样问题导致没有基质像素怎么办，即jizhi_position_list为空，这种情况下要单独处理
    if len(jizhi_position_list) == 0:
        jizhi_position_list = []
        jizhi_activations = []
        print('基质元素个数为0')
    else:
        # 去除基质位置列表与基底膜位置列表中重复的元素
        jizhi_position_list = list(set(jizhi_position_list) - set(jidimo_position_list))
        # 定义基质激活值的上界upper_bound和下界lower_bound
        jizhi_activation_upper_bound = 0.95
        jizhi_activation_lower_bound = 0.3
        jizhi_dis_activations = get_hm_activation_according_to_distances(jidimo_position_list, jizhi_position_list,
                                                                         jizhi_activation_upper_bound,
                                                                         jizhi_activation_lower_bound)
        # 融合基质的灰度值重新计算激活值
        jizhi_activations = fuse_gray_scale_activations(image, jizhi_position_list, jizhi_dis_activations)
        print('基质激活值已经计算完成')

    # 接下来计算保护套对应的每一个像素的位置的热力图激活值
    # 保护套位置像素的每个位置的激活值都为1e-3
    baohutao_position_list = find_pos_np(mask, 4)
    # 这里就存在一个问题，如果一个图片中因为采样问题导致没有保护套像素怎么办，即baohutao_position_list为空，这种情况下要单独处理
    if len(baohutao_position_list) == 0:
        baohutao_position_list = []
        baohutao_activation_list = []
        print('保护套元素个数为0')
    else:
        baohutao_activation = 1e-3
        baohutao_activation_list = [baohutao_activation] * len(baohutao_position_list)
        print('保护套激活值已经计算完成')

    labeled_position_list = []
    labeled_position_list.extend(jidimo_position_list)
    labeled_position_list.extend(shangpi_position_list)
    labeled_position_list.extend(jizhi_position_list)
    labeled_position_list.extend(baohutao_position_list)

    labeled_activation_list = []
    labeled_activation_list.extend(jidimo_activation_list)
    labeled_activation_list.extend(shangpi_activations)
    labeled_activation_list.extend(jizhi_activations)
    labeled_activation_list.extend(baohutao_activation_list)

    return labeled_position_list, labeled_activation_list


def generate_yan_heatmap_activation(image, mask):
    labeled_position_list, labeled_activation_list = generate_zhengchang_heatmap_activation(image, mask)
    # 其他位置的热力图激活值都为0，比如背景或其他未标记的位置
    all_position_list = get_coordinates(mask)
    other_position_list = list(set(all_position_list) - set(labeled_position_list))
    other_position_activations = [0.0 for i in range(len(other_position_list))]

    # 构建所有位置到其计算得到的激活值的字典
    pos2activation = defaultdict(float)
    for i in range(len(labeled_position_list)):
        pos2activation[labeled_position_list[i]] = labeled_activation_list[i]
    for i in range(len(other_position_list)):
        pos2activation[other_position_list[i]] = other_position_activations[i]

    # 定义二维数组的初始激活值
    H, W = mask.shape[0], mask.shape[1]
    heatmap = np.zeros((H, W))
    for i in range(1, H - 1):
        for j in range(1, W - 1):
            if isinstance(pos2activation[(i, j)], Iterable):
                pos2activation[(i, j)] = pos2activation[(i, j)][0]
            heatmap[i, j] = pos2activation[(i, j)]
    # 对于上下左右四个边界的激活值，我们将其设置为0，进行32倍下采样时，边界的0值可以忽略不计

    return heatmap


def generate_nangzhong_heatmap_activation(image, mask):
    """
    这个函数用来根据囊肿图片的分割掩码生成自定义的热力图激活
    :param image: 3d array, [H, W, 3]
    :param mask: 2d array, [H, W]
    :return: 热力图激活叠加原图的3d array, [H, W, 3]，可以直接展示
    """
    # 首先识别出囊肿的边缘，当当前像素为囊肿，且其上下左右四个像素中有一个像素不为囊肿，则该
    # 像素属于囊肿的边缘(轮廓)
    H, W = mask.shape[0], mask.shape[1]
    nangzhong_lunkuo_position_list = []
    for i in range(1, H - 1):
        for j in range(1, W - 1):
            if mask[i, j] == 6:
                if mask[i, j - 1] != 6 or mask[i, j + 1] != 6 or mask[i - 1, j] != 6 or mask[i + 1, j] != 6:
                    nangzhong_lunkuo_position_list.append((i, j))

    print('囊肿轮廓激活值已经计算完成')
    # 定义囊肿位置列表，用来保存
    nangzhong_position_list = find_pos_np(mask, 6)
    # 去除囊肿位置列表与囊肿轮廓位置列表中重复的元素，获得囊肿内部位置列表，用来保存
    nangzhong_neibu_position_list = list(set(nangzhong_position_list) - set(nangzhong_lunkuo_position_list))
    print('囊肿内部激活值已经计算完成')

    # 定义囊肿轮廓的激活值和囊肿内部的激活值
    nz_lk_activation = 1.0
    nz_nb_activation = 0.8

    # 下面的就跟炎症的流程是一样的, 顺序分别为基底膜，上皮，基质，保护套
    labeled_position_list, labeled_activation_list = generate_zhengchang_heatmap_activation(image, mask)
    labeled_position_list.extend(nangzhong_lunkuo_position_list)
    labeled_position_list.extend(nangzhong_neibu_position_list)
    labeled_activation_list.extend([nz_lk_activation] * len(nangzhong_lunkuo_position_list))
    labeled_activation_list.extend([nz_nb_activation] * len(nangzhong_neibu_position_list))

    # 其他位置的热力图激活值都为0，比如背景或其他未标记的位置
    all_position_list = get_coordinates(mask)
    other_position_list = list(set(all_position_list) - set(labeled_position_list))
    other_position_activations = [0.0 for i in range(len(other_position_list))]

    # 构建所有位置到其计算得到的激活值的字典
    pos2activation = defaultdict(float)
    for i in range(len(labeled_position_list)):
        pos2activation[labeled_position_list[i]] = labeled_activation_list[i]
    for i in range(len(other_position_list)):
        pos2activation[other_position_list[i]] = other_position_activations[i]
    # 定义二维数组的初始激活值
    H, W = mask.shape[0], mask.shape[1]
    heatmap = np.zeros((H, W))
    for i in range(1, H - 1):
        for j in range(1, W - 1):
            if isinstance(pos2activation[(i, j)], Iterable):
                pos2activation[(i, j)] = pos2activation[(i, j)][0]
            heatmap[i, j] = pos2activation[(i, j)]
    # 对于上下左右四个边界的激活值，我们将其设置为0，进行32倍下采样时，边界的0值可以忽略不计

    return heatmap


def generate_waifan_heatmap_activation(image, mask):
    """
    这个函数用来根据外翻图片的分割掩码生成自定义的热力图激活
    :param image: 3d array, [H, W, 3]
    :param mask: 2d array, [H, W]
    :return: 热力图激活叠加原图的3d array, [H, W, 3]，可以直接展示
    """
    image = image[:, :, 0]  # 取一个channel变为灰度图
    H, W = mask.shape[0], mask.shape[1]
    # 首先识别出凸起的边缘，当当前像素为凸起，且其上下左右四个像素中有一个像素不为凸起，则该
    #  像素为凸起边缘，记录其位置
    tuqi_lunkuo_position_list = []
    for i in range(1, H - 1):
        for j in range(1, W - 1):
            if mask[i, j] == 5:
                if mask[i, j - 1] != 5 or mask[i, j + 1] != 5 or mask[i - 1, j] != 5 or mask[i + 1, j] != 5:
                    tuqi_lunkuo_position_list.append((i, j))
    # 定义凸起轮廓的激活值
    tuqi_lk_activation = 1.0
    tuqi_lk_activations = [tuqi_lk_activation] * len(tuqi_lunkuo_position_list)
    print('凸起轮廓激活值已经计算完成')
    # 定义凸起位置列表，用来保存
    tuqi_position_list = find_pos_np(mask, 5)
    # 去除凸起位置列表与凸起轮廓位置列表中重复的元素，获得凸起内部位置列表，用来保存
    tuqi_neibu_position_list = list(set(tuqi_position_list) - set(tuqi_lunkuo_position_list))

    # 定义凸起激活值的上界upper_bound和下界lower_bound
    tuqi_activation_upper_bound = 1.0
    tuqi_activation_lower_bound = 0.1
    if len(tuqi_lunkuo_position_list) == 0 and len(tuqi_neibu_position_list) == 0:
        tuqi_lunkuo_position_list = []
        tuqi_neibu_position_list = []
        tuqi_lk_activations = []
        tuqi_neibu_activations = []
        print('凸起元素个数为0')
    else:
        tuqi_neibu_dis_activations = get_hm_activation_according_to_distances(tuqi_lunkuo_position_list,
                                                                              tuqi_neibu_position_list,
                                                                              tuqi_activation_upper_bound,
                                                                              tuqi_activation_lower_bound)
        # 定义滤波器的大小
        kernel_size = (7, 7)
        # 对数组进行平均滤波
        image = cv2.blur(image, kernel_size)
        # 融合凸起的灰度值重新计算激活值
        tuqi_neibu_activations = fuse_gray_scale_activations(image, tuqi_neibu_position_list, tuqi_neibu_dis_activations)
        print('凸起激活值已经计算完成')

    # 接下来计算保护套对应的每一个像素的位置的热力图激活值
    # 保护套位置像素的每个位置的激活值都为1e-3
    baohutao_position_list = find_pos_np(mask, 4)
    # 这里就存在一个问题，如果一个图片中因为采样问题导致没有保护套像素怎么办，即baohutao_position_list为空，这种情况下要单独处理
    if len(baohutao_position_list) == 0:
        baohutao_position_list = []
        baohutao_activation_list = []
        print('保护套元素个数为0')
    else:
        baohutao_activation = 1e-3
        baohutao_activation_list = [baohutao_activation] * len(baohutao_position_list)
        print('保护套激活值已经计算完成')

    # 下面的就跟炎症的流程是一样的, 顺序分别为基底膜，上皮，基质，保护套
    labeled_position_list, labeled_activation_list = [], []
    labeled_position_list.extend(tuqi_lunkuo_position_list)
    labeled_position_list.extend(tuqi_neibu_position_list)
    labeled_position_list.extend(baohutao_position_list)

    labeled_activation_list.extend(tuqi_lk_activations)
    labeled_activation_list.extend(tuqi_neibu_activations)
    labeled_activation_list.extend(baohutao_activation_list)

    # 其他位置的热力图激活值都为0，比如背景或其他未标记的位置
    all_position_list = get_coordinates(mask)
    other_position_list = list(set(all_position_list) - set(labeled_position_list))
    other_position_activations = [0.0 for i in range(len(other_position_list))]

    # 构建所有位置到其计算得到的激活值的字典
    pos2activation = defaultdict(float)
    for i in range(len(labeled_position_list)):
        pos2activation[labeled_position_list[i]] = labeled_activation_list[i]
    for i in range(len(other_position_list)):
        pos2activation[other_position_list[i]] = other_position_activations[i]

    # 定义二维数组的初始激活值
    H, W = mask.shape[0], mask.shape[1]
    heatmap = np.zeros((H, W))
    for i in range(1, H - 1):
        for j in range(1, W - 1):
            if isinstance(pos2activation[(i, j)], Iterable):
                pos2activation[(i, j)] = pos2activation[(i, j)][0]
            heatmap[i, j] = pos2activation[(i, j)]
    # 对于上下左右四个边界的激活值，我们将其设置为0，进行32倍下采样时，边界的0值可以忽略不计

    return heatmap


def parse_xml(xml_file_path):
    # 解析XML文件
    tree = ET.parse(xml_file_path)
    root = tree.getroot()

    # 提取 'object' 信息
    objects = []
    for obj in root.findall('object'):
        object_info = {
            'name': obj.find('name').text,
            'bndbox': {
                'xmin': int(obj.find('bndbox/xmin').text),
                'ymin': int(obj.find('bndbox/ymin').text),
                'xmax': int(obj.find('bndbox/xmax').text),
                'ymax': int(obj.find('bndbox/ymax').text),
            }
        }
        objects.append(object_info)

    # 打印提取的对象信息
    for object_info in objects:
        print(object_info)

    return objects


# def generate_hsil_heatmap_activation(image, xml_file_path, worksheet):
#     """
#     这个函数用来根据高级别病变的XML文件中的标注的bounding box生成自定义的热力图激活
#     :param image: 3d array, [H, W, 3]
#     :param xml_file_path: 图片对应的xml文件
#     :param worksheet: excel对象，用于记录没有标记的图片的路径
#     :return: 热力图激活叠加原图的3d array, [H, W, 3]，可以直接展示
#     """
#     image = image[:, :, 0]  # 取一个channel变为灰度图
#     H, W = image.shape[0], image.shape[1]
#     # 初始化热力图
#     heatmap = np.zeros((H, W))
#     lower_bound = 0.2
#     linear_func = linear_function_through_points((0.0, 1.0),
#                                                  (255.0, lower_bound))
#     try:
#         objects = parse_xml(xml_file_path)
#         for obj in objects:
#             bb_type = obj['name']
#             bb_xmin = obj['bndbox']['xmin']
#             bb_ymin = obj['bndbox']['ymin']
#             bb_xmax = obj['bndbox']['xmax']
#             bb_ymax = obj['bndbox']['ymax']
#             # 取出图像中边框盒内部的像素值
#             img_bb = image[bb_ymin:bb_ymax, bb_xmin:bb_xmax]
#             # 定义滤波器的大小
#             kernel_size = (7, 7)
#             # 对数组进行平均滤波
#             img_bb = cv2.blur(img_bb, kernel_size)
#             for i in range(bb_ymin, bb_ymax):
#                 for j in range(bb_xmin, bb_xmax):
#                     heatmap[i, j] = linear_func(img_bb[i - bb_ymin, j - bb_xmin])
#
#             return heatmap
#     except Exception:
#         worksheet.append([xml_file_path])
#         print(xml_file_path, 'is not labeled')
#         return None



def generate_hsil_heatmap_activation(image, xml_file_path):
    """
    这个函数用来根据高级别病变的XML文件中的标注的bounding box生成自定义的热力图激活
    :param image: 3d array, [H, W, 3]
    :param xml_file_path: 图片对应的xml文件
    :return: 热力图激活叠加原图的3d array, [H, W, 3]，可以直接展示
    """
    image = image[:, :, 0]  # 取一个channel变为灰度图
    H, W = image.shape[0], image.shape[1]
    # 初始化热力图
    heatmap = np.zeros((H, W))

    lower_bound = 0.2
    linear_func = linear_function_through_points((0.0, 1.0),
                                                 (255.0, lower_bound))

    objects = parse_xml(xml_file_path)
    for obj in objects:
        bb_type = obj['name']
        bb_xmin = obj['bndbox']['xmin']
        bb_ymin = obj['bndbox']['ymin']
        bb_xmax = obj['bndbox']['xmax']
        bb_ymax = obj['bndbox']['ymax']
        # 取出图像中边框盒内部的像素值
        img_bb = image[bb_ymin:bb_ymax, bb_xmin:bb_xmax]
        # 定义滤波器的大小
        kernel_size = (7, 7)
        # 对数组进行平均滤波
        img_bb = cv2.blur(img_bb, kernel_size)
        for i in range(bb_ymin, bb_ymax):
            for j in range(bb_xmin, bb_xmax):
                heatmap[i, j] = linear_func(img_bb[i - bb_ymin, j - bb_xmin])
        return heatmap




def get_image_heatmap(img, heatmap):
    """
    这个函数用来将自定义的热力图激活叠加原图的3d array, [H, W, 3]，可以直接展示
    :param img: 3d array, [H, W, 3]
    :param heatmap: 3d array, [H, W]
    :return:
    """
    visualization = show_cam_on_image(img / 255., heatmap, use_rgb=True)
    # 在保存之前，将图像从RGB转换为BGR
    img_bgr = cv2.cvtColor(visualization, cv2.COLOR_RGB2BGR)
    return img_bgr


def get_all_heatmaps():
    """
    该函数用来根据分割掩码或bounding box生成所有的热力图
    :return:
    """
    # 创建一个excel表格用于存储预测结果
    workbook = openpyxl.Workbook()
    # 获取当前活跃的worksheet，默认就是第一个worksheet
    worksheet = workbook.active
    worksheet.title = 'file'
    # 定义所要输出的文件的表头
    header = ['文件名', '分割掩码路径']
    worksheet.append(header)

    file_path = os.path.join('dataset_seg', 'merge_seg.xlsx')
    # 读取结果的excel文件
    df = pd.read_excel(io=file_path, sheet_name='file')
    num_rows = len(df)
    # df = df[: num_rows // 2]
    df = df[7494: num_rows // 2]
    for idx, row in df.iterrows():
        # ['文件名', '分割掩码路径']
        img_path = row[0]
        mask_path = row[1]
        img_type = img_path.split('/')[-2]
        # 读取原图和分割掩码图
        start_time = time.time()

        img = cv2.imread(img_path)
        mask = cv2.imread(mask_path, cv2.IMREAD_GRAYSCALE)
        # 根据图片类别调用不同的获取激活图的函数
        if img_type == '宫颈炎':
            activation_map = generate_yan_heatmap_activation(img, mask)
        elif img_type == '囊肿':
            activation_map = generate_nangzhong_heatmap_activation(img, mask)
        elif img_type == '外翻':
            activation_map = generate_waifan_heatmap_activation(img, mask)
        # 定义生成的激活图的序列化的存储路径
        path_list = img_path.split('/')
        path_list_copy = deepcopy(path_list)
        path_list.insert(-2, 'activation_map')
        activation_map_path = '/'.join(path_list).replace('.png', '.npy')
        directory = os.path.dirname(activation_map_path)
        os.makedirs(directory, exist_ok=True)
        np.save(activation_map_path, activation_map)

        # 根据原图和激活图生成热力图
        heatmap = get_image_heatmap(img, activation_map)
        path_list_copy.insert(-2, 'heatmap')
        heatmap_path = '/'.join(path_list_copy)
        heatmap_directory = os.path.dirname(heatmap_path)
        os.makedirs(heatmap_directory, exist_ok=True)
        cv2.imwrite(heatmap_path, heatmap)
        worksheet.append([img_path, activation_map_path,  heatmap_path])
        print(str(idx + 1), img_path, activation_map_path, heatmap_path)
        # 记录结束时间
        end_time = time.time()
        # 计算并打印运行时间
        print("程序运行时间：", end_time - start_time, "秒")

    workbook.save(filename=os.path.join('dataset_seg', 'merge_heatmap0.xlsx'))


def get_all_heatmaps_for_hsil():
    """
    该函数用来根据bounding box生成所有的激活图
    :return:
    """
    # 创建一个excel表格用于存储预测结果
    workbook = openpyxl.Workbook()
    # 获取当前活跃的worksheet，默认就是第一个worksheet
    worksheet = workbook.active
    worksheet.title = 'file'
    # 定义所要输出的文件的表头
    header = ['文件名', '激活图路径', '热力图路径']
    worksheet.append(header)

    # 创建一个excel表格用于存储没有标注的图像文件
    workbook1 = openpyxl.Workbook()
    # 获取当前活跃的worksheet，默认就是第一个worksheet
    worksheet1 = workbook1.active
    worksheet1.title = 'file'
    # 定义所要输出的文件的表头
    header1 = ['文件名']
    worksheet1.append(header1)

    file_path = os.path.join('dataset', 'tiff_xiangya_frame.xlsx')
    # 读取结果的excel文件
    df = pd.read_excel(io=file_path, sheet_name='file')
    df = df[df['类别'].isin([3, 4])]
    num_rows = len(df)
    df = df[: num_rows // 2]
    df = df.reset_index()
    print('-'*100)
    print(len(df))
    print('-'*100)
    for idx, row in df.iterrows():
        # ['文件名', '分割掩码路径']
        img_path = row['文件名']
        img_type = int(row['类别'])
        # 读取原图和xml标注文件
        start_time = time.time()
        img = cv2.imread(img_path)
        # 根据图片的存储路径推导出xml的文件路径
        path_list = img_path.split('/')
        path_list_copy = deepcopy(path_list)
        path_list_copy_1 = deepcopy(path_list)
        path_list.insert(-2, 'segmentation_mask')
        xml_file_path = '/'.join(path_list).replace('.png', '.xml')
        activation_map = generate_hsil_heatmap_activation(img, xml_file_path, worksheet1)
        if activation_map is None:
            continue

        # 定义生成的激活图的序列化的存储路径
        path_list_copy.insert(-2, 'activation_map')
        activation_map_path = '/'.join(path_list_copy).replace('.png', '.npy')
        directory = os.path.dirname(activation_map_path)
        os.makedirs(directory, exist_ok=True)
        np.save(activation_map_path, activation_map)

        # 根据原图和激活图生成热力图
        heatmap = get_image_heatmap(img, activation_map)
        path_list_copy_1.insert(-2, 'heatmap')
        heatmap_path = '/'.join(path_list_copy_1)
        heatmap_directory = os.path.dirname(heatmap_path)
        os.makedirs(heatmap_directory, exist_ok=True)
        cv2.imwrite(heatmap_path, heatmap)
        worksheet.append([img_path, activation_map_path, heatmap_path])
        print(str(idx + 1), img_path, activation_map_path, heatmap_path)
        # 记录结束时间
        end_time = time.time()
        # 计算并打印运行时间
        print("程序运行时间：", end_time - start_time, "秒")

    workbook.save(filename=os.path.join('dataset_seg', 'hsil_heatmap4.xlsx'))
    workbook1.save(filename=os.path.join('dataset_seg', 'hsil_error4.xlsx'))


if __name__ == '__main__':
    # get_all_heatmaps()
    # get_all_heatmaps_for_hsil()
    # 记录开始时间
    # start_time = time.time()
    #
    img = cv2.imread('segmentation_mask/image/囊肿/M0008_2021_P0000535_circle_2.0x2.5_C12_S12_6.png')
    mask = cv2.imread('segmentation_mask/annotation/囊肿/M0008_2021_P0000535_circle_2.0x2.5_C12_S12_6.png',
                      cv2.IMREAD_GRAYSCALE)
    heatmap = generate_nangzhong_heatmap_activation(img, mask)
    # for i in range(heatmap.shape[0]):
    #     print(heatmap[i])
    visualization = show_cam_on_image(img / 255., heatmap, use_rgb=True)
    # 在保存之前，将图像从RGB转换为BGR
    img_bgr = cv2.cvtColor(visualization, cv2.COLOR_RGB2BGR)
    cv2.imwrite('yan_heatmap_activation.png', img_bgr)
    plt.imshow(visualization)
    plt.show()
    #
    # # 记录结束时间
    # end_time = time.time()
    # # 计算并打印运行时间
    # print("程序运行时间：", end_time - start_time, "秒")

    # img_file_path = 'segmentation_mask/image/宫颈癌/M0008_2021_P0000246_circle_2.4x3.0_C4_S4_0.png'
    # xml_file_path = 'segmentation_mask/annotation/宫颈癌/M0008_2021_P0000246_circle_2.4x3.0_C4_S4_0.xml'
    # img = cv2.imread(img_file_path)
    # # parse_xml(xml_file_path)
    # heatmap = generate_hsil_heatmap_activation(img, xml_file_path)
    # visualization = show_cam_on_image(img / 255., heatmap, use_rgb=True)
    # # 在保存之前，将图像从RGB转换为BGR
    # img_bgr = cv2.cvtColor(visualization, cv2.COLOR_RGB2BGR)
    # cv2.imwrite('yan_heatmap_activation.png', img_bgr)
    # plt.imshow(visualization)
    # plt.show()
